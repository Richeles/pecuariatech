import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Cria uma nova planilha
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Financeiro"

# Cabeçalho
headers = ["descricao", "tipo", "valor", "categoria", "data_lancamento"]
ws.append(headers)

# Dados de exemplo
dados = [
    ["Venda de Novilhos", "receita", 8500, "Venda", "2026-06-15"],
    ["Venda de Bezerros", "receita", 3200, "Venda", "2026-06-20"],
    ["Compra de Sal Mineral", "despesa", 450, "Alimentação", "2026-06-10"],
    ["Vacinação do Rebanho", "despesa", 780, "Sanidade", "2026-06-12"],
    ["Mão de Obra", "despesa", 1200, "Operacional", "2026-06-18"],
]

for linha in dados:
    ws.append(linha)

# Formatação básica
for coluna in range(1, 6):
    celula = ws.cell(row=1, column=coluna)
    celula.font = Font(bold=True)
    celula.alignment = Alignment(horizontal="center")
    celula.fill = PatternFill(start_color="d3d3d3", end_color="d3d3d3", fill_type="solid")

# Ajusta largura das colunas
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = max_length + 2
    ws.column_dimensions[col_letter].width = adjusted_width

# Salva o arquivo
nome_arquivo = "teste_financeiro.xlsx"
wb.save(nome_arquivo)
print(f"✅ Arquivo '{nome_arquivo}' criado com sucesso!")