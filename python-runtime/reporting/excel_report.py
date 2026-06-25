import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from dto.dashboard_dto import DashboardDTO

def gerar_excel(dto: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    # Cabecalhos
    headers = ['Indicador', 'Valor']
    ws.append(headers)
    for col in ['A', 'B']:
        ws[col + '1'].font = Font(bold=True)
        ws[col + '1'].fill = PatternFill(start_color="A9D08E", fill_type="solid")
        ws[col + '1'].alignment = Alignment(horizontal='center')

    data = [
        ['Score pi', dto.get('score_pi', 0)],
        ['Capital Score', dto.get('capital_score', 0)],
        ['Risco Estrutural', dto.get('risco_estrutural', 'N/A')],
        ['ROI (%)', dto.get('roi', 0)],
        ['Margem (%)', dto.get('margem', 0)],
        ['EBITDA (R$)', dto.get('ebitda', 0)],
        ['GMD (kg/dia)', dto.get('gmd', 0)],
        ['Lotacao (UA/ha)', dto.get('lotacao', 0)],
        ['Governanca', dto.get('governanca', 0)],
        ['ESG', dto.get('esg', 0)],
        ['Rastreabilidade', dto.get('rastreabilidade', 0)],
        ['Maturidade Digital', dto.get('maturidade_digital', 0)],
        ['Compliance', dto.get('compliance', 0)],
        ['Capital Intelectual', dto.get('capital_intelectual', 0)],
    ]

    for row in data:
        ws.append(row)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()